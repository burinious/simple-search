import json
import os
import re
import sys
from datetime import datetime

from openpyxl import load_workbook


DEFAULT_SOURCES = [
    r"C:\Users\USER PCX\Downloads\Without_ClassOfDegree.xlsx",
    r"C:\Users\USER PCX\Downloads\With_ClassOfDegree.xlsx",
]

KEEP_FIELDS = [
    "MatricNo",
    "FirstName",
    "Middlename",
    "Surname",
    "GSMNo",
    "StateOfOrigin",
    "ClassOfDegree",
    "DateOfBirth",
    "DateOfGraduation",
    "Status",
    "Gender",
    "MaritalStatus",
    "JambRegNo",
    "IsMilitary",
    "CourseOfStudy",
    "StudyMode",
]

NAME_FIELDS = {"FirstName", "Middlename", "Surname"}
TITLE_FIELDS = {
    "StateOfOrigin",
    "CourseOfStudy",
    "MaritalStatus",
    "StudyMode",
    "ClassOfDegree",
    "Status",
}


def normalize_text(field, value):
    text = "" if value is None else str(value).strip()
    if not text:
        return ""
    if field in NAME_FIELDS or field in TITLE_FIELDS:
        return text.title()
    if field in {"Gender", "IsMilitary"}:
        return text.upper()
    return text


def normalize_date(value):
    if value in (None, ""):
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y")

    text = str(value).strip()
    if not text:
        return ""

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue

    return text


def load_students(source_paths):
    students = []

    for path in source_paths:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        header = ["" if value is None else str(value).strip() for value in next(iterator)]
        indices = {name: header.index(name) for name in KEEP_FIELDS}

        for row in iterator:
            student = {}
            for field in KEEP_FIELDS:
                raw_value = row[indices[field]]
                if field in {"DateOfBirth", "DateOfGraduation"}:
                    student[field] = normalize_date(raw_value)
                else:
                    student[field] = normalize_text(field, raw_value)
            students.append(student)

    students.sort(
        key=lambda item: (
            item["Surname"].lower(),
            item["FirstName"].lower(),
            item["Middlename"].lower(),
            item["MatricNo"].lower(),
        )
    )
    return students


def write_students_json(students, output_path):
    with open(output_path, "w", encoding="utf-8") as handle:
        json.dump(students, handle, indent=2, ensure_ascii=False)


def write_search_files(students, index_path, records_dir):
    os.makedirs(records_dir, exist_ok=True)

    for name in os.listdir(records_dir):
        path = os.path.join(records_dir, name)
        if os.path.isfile(path):
            os.remove(path)

    index = []
    for student in students:
        matric = student["MatricNo"]
        safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", matric) or "record"
        filename = f"{safe_name}.json"
        full_name = " ".join(
            part
            for part in [student["Surname"], student["FirstName"], student["Middlename"]]
            if part
        ).strip()

        with open(os.path.join(records_dir, filename), "w", encoding="utf-8") as handle:
            json.dump(student, handle, indent=2, ensure_ascii=False)

        index.append(
            {
                "MatricNo": matric,
                "FirstName": student["FirstName"],
                "Middlename": student["Middlename"],
                "Surname": student["Surname"],
                "FullName": full_name,
                "RecordFile": f"{records_dir}/{filename}",
            }
        )

    with open(index_path, "w", encoding="utf-8") as handle:
        json.dump(index, handle, indent=2, ensure_ascii=False)


def main():
    source_paths = sys.argv[1:] or DEFAULT_SOURCES

    missing = [path for path in source_paths if not os.path.exists(path)]
    if missing:
        for path in missing:
            print(f"Missing file: {path}")
        raise SystemExit(1)

    students = load_students(source_paths)
    write_students_json(students, "students.json")
    write_search_files(students, "search-index.json", "student-records")

    print(f"Updated {len(students)} student records.")
    print("Wrote students.json, search-index.json, and student-records/*.json")


if __name__ == "__main__":
    main()
